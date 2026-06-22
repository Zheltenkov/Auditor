"""Оценка качества аудита на корпусе проектов с Excel-разметкой."""

from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from content_audit.domain import AuditReport, CRITERION_LABELS, Criterion


PROJECT_COLUMN = "Проект"
PROBLEM_COLUMN = "Проблема"
DETAILS_COLUMN = "Детали"


CRITERION_ALIASES: dict[Criterion, tuple[str, ...]] = {
    Criterion.ACTUALITY: (
        "актуаль",
        "неактуаль",
        "устар",
        "ссылка",
        "версии ос",
        "инструмент",
        "api",
        "сервис",
    ),
    Criterion.READABILITY: (
        "опечат",
        "граммат",
        "орфограф",
        "формулиров",
        "нумерац",
        "двоеточ",
    ),
    Criterion.CHECKLIST_ALIGNMENT: (
        "чек-лист",
        "чеклист",
        "check-list",
        "несоответствие задания чек-листу",
        "несоответствие задания чеклисту",
    ),
    Criterion.CORRECTNESS: (
        "ошибка в задании",
        "нарушение логики",
        "логик",
        "противореч",
        "некоррект",
        "неверн",
        "ошибка в доп материалах",
        "ошибка в дополнительных материалах",
    ),
}


class CorpusEvaluationKey(BaseModel, frozen=True):
    """Ключ сравнения: один проект и один критерий."""

    project_id: str
    criterion: str


class GoldCorpusItem(BaseModel):
    """Одна эталонная строка после нормализации Excel."""

    row_number: int
    raw_project: str
    matched_project: str
    project_id: str
    raw_problem: str
    details: str
    criteria: list[str]


class CriterionMetrics(BaseModel):
    """Метрики по одному критерию."""

    criterion: str
    label: str
    gold_total: int
    predicted_total: int
    true_positive: int
    false_positive: int
    false_negative: int
    precision: float
    recall: float
    f1_score: float


class CorpusEvaluationSummary(BaseModel):
    """Итог оценки по корпусу проектов."""

    evaluated_criteria: list[str]
    gold_total: int
    predicted_total: int
    true_positive: int
    false_positive: int
    false_negative: int
    precision: float
    recall: float
    f1_score: float
    macro_precision: float
    macro_recall: float
    macro_f1_score: float
    gold_scope_predicted_total: int
    gold_scope_true_positive: int
    gold_scope_false_positive: int
    gold_scope_false_negative: int
    gold_scope_precision: float
    gold_scope_recall: float
    gold_scope_f1_score: float
    gold_scope_macro_precision: float
    gold_scope_macro_recall: float
    gold_scope_macro_f1_score: float
    per_criterion: list[CriterionMetrics]
    gold_items: list[GoldCorpusItem]
    false_positive_items: list[CorpusEvaluationKey]
    false_negative_items: list[CorpusEvaluationKey]
    project_mapping: dict[str, str]
    notes: list[str] = Field(default_factory=list)


@dataclass(frozen=True)
class _ProjectCandidate:
    project_id: str
    raw_name: str
    normalized_name: str
    tokens: frozenset[str]


def evaluate_corpus_report(report: AuditReport, gold_xlsx_path: Path) -> CorpusEvaluationSummary:
    """Сравнивает отчёт аудита с Excel-разметкой на уровне `проект × критерий`."""

    unit_candidates = _project_candidates_from_report(report)
    gold_items, mapping_notes = load_gold_items(gold_xlsx_path, unit_candidates)
    gold_keys = {
        CorpusEvaluationKey(project_id=item.project_id, criterion=criterion)
        for item in gold_items
        for criterion in item.criteria
    }
    predicted_keys = _predicted_keys_from_report(report)

    true_positive_keys = gold_keys & predicted_keys
    false_positive_keys = predicted_keys - gold_keys
    false_negative_keys = gold_keys - predicted_keys
    per_criterion = _per_criterion_metrics(gold_keys, predicted_keys)
    evaluated_criteria = sorted({item.criterion for item in gold_keys})
    gold_scope_predicted_keys = {item for item in predicted_keys if item.criterion in evaluated_criteria}
    gold_scope_true_positive_keys = gold_keys & gold_scope_predicted_keys
    gold_scope_false_positive_keys = gold_scope_predicted_keys - gold_keys
    gold_scope_false_negative_keys = gold_keys - gold_scope_predicted_keys
    gold_scope_metrics = [item for item in per_criterion if item.criterion in evaluated_criteria]

    return CorpusEvaluationSummary(
        evaluated_criteria=evaluated_criteria,
        gold_total=len(gold_keys),
        predicted_total=len(predicted_keys),
        true_positive=len(true_positive_keys),
        false_positive=len(false_positive_keys),
        false_negative=len(false_negative_keys),
        precision=_safe_ratio(len(true_positive_keys), len(true_positive_keys) + len(false_positive_keys)),
        recall=_safe_ratio(len(true_positive_keys), len(true_positive_keys) + len(false_negative_keys)),
        f1_score=_f1(len(true_positive_keys), len(false_positive_keys), len(false_negative_keys)),
        macro_precision=_mean([item.precision for item in per_criterion]),
        macro_recall=_mean([item.recall for item in per_criterion]),
        macro_f1_score=_mean([item.f1_score for item in per_criterion]),
        gold_scope_predicted_total=len(gold_scope_predicted_keys),
        gold_scope_true_positive=len(gold_scope_true_positive_keys),
        gold_scope_false_positive=len(gold_scope_false_positive_keys),
        gold_scope_false_negative=len(gold_scope_false_negative_keys),
        gold_scope_precision=_safe_ratio(
            len(gold_scope_true_positive_keys),
            len(gold_scope_true_positive_keys) + len(gold_scope_false_positive_keys),
        ),
        gold_scope_recall=_safe_ratio(
            len(gold_scope_true_positive_keys),
            len(gold_scope_true_positive_keys) + len(gold_scope_false_negative_keys),
        ),
        gold_scope_f1_score=_f1(
            len(gold_scope_true_positive_keys),
            len(gold_scope_false_positive_keys),
            len(gold_scope_false_negative_keys),
        ),
        gold_scope_macro_precision=_mean([item.precision for item in gold_scope_metrics]),
        gold_scope_macro_recall=_mean([item.recall for item in gold_scope_metrics]),
        gold_scope_macro_f1_score=_mean([item.f1_score for item in gold_scope_metrics]),
        per_criterion=per_criterion,
        gold_items=gold_items,
        false_positive_items=sorted(false_positive_keys, key=lambda item: (item.project_id, item.criterion)),
        false_negative_items=sorted(false_negative_keys, key=lambda item: (item.project_id, item.criterion)),
        project_mapping={item.raw_project: item.matched_project for item in gold_items},
        notes=[
            "Сравнение выполняется на уровне проект × критерий, без сравнения строк и цитат.",
            "Excel-разметка нормализуется эвристически из колонок 'Проблема' и 'Детали'.",
            *mapping_notes,
        ],
    )


def load_gold_items(
    gold_xlsx_path: Path,
    unit_candidates: list[_ProjectCandidate],
) -> tuple[list[GoldCorpusItem], list[str]]:
    """Читает Excel и переводит строки в эталонные критерии."""

    workbook = load_workbook(gold_xlsx_path, data_only=True)
    sheet = workbook.active
    header = _header_map(sheet)
    required_columns = {PROJECT_COLUMN, PROBLEM_COLUMN, DETAILS_COLUMN}
    missing_columns = sorted(required_columns - set(header))
    if missing_columns:
        raise ValueError(f"В Excel не найдены обязательные колонки: {', '.join(missing_columns)}")

    items: list[GoldCorpusItem] = []
    notes: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_project = _cell_text(row, header[PROJECT_COLUMN])
        raw_problem = _cell_text(row, header[PROBLEM_COLUMN])
        details = _cell_text(row, header[DETAILS_COLUMN])
        if not raw_project and not raw_problem and not details:
            continue

        candidate, score = _match_project(raw_project, unit_candidates)
        criteria = _criteria_from_gold_row(raw_problem, details)
        if not criteria:
            notes.append(f"Строка {row_number}: не удалось вывести критерий из типа проблемы {raw_problem!r}.")
            continue
        if score < 0.55:
            notes.append(
                f"Строка {row_number}: слабое сопоставление проекта {raw_project!r} "
                f"с папкой {candidate.raw_name!r}, score={score:.2f}."
            )

        items.append(
            GoldCorpusItem(
                row_number=row_number,
                raw_project=raw_project,
                matched_project=candidate.raw_name,
                project_id=candidate.project_id,
                raw_problem=raw_problem,
                details=details,
                criteria=criteria,
            )
        )
    return items, notes


def write_corpus_evaluation(summary: CorpusEvaluationSummary, output_dir: Path) -> None:
    """Записывает машинный и табличный отчёт оценки."""

    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "corpus_evaluation.json").write_text(
        json.dumps(summary.model_dump(mode="json"), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _write_metrics_csv(summary, output_dir / "corpus_evaluation_by_criterion.csv")
    _write_items_csv(summary.false_negative_items, output_dir / "corpus_false_negative.csv")
    _write_items_csv(summary.false_positive_items, output_dir / "corpus_false_positive.csv")


def _project_candidates_from_report(report: AuditReport) -> list[_ProjectCandidate]:
    """Создаёт кандидаты сопоставления из единиц отчёта."""

    return [
        _ProjectCandidate(
            project_id=unit.unit_id,
            raw_name=unit.name,
            normalized_name=_normalize_project_name(unit.name),
            tokens=frozenset(_project_tokens(unit.name)),
        )
        for unit in report.units
    ]


def _predicted_keys_from_report(report: AuditReport) -> set[CorpusEvaluationKey]:
    """Берёт все критерии, по которым алгоритм нашёл хотя бы один случай."""

    unit_ids = {unit.unit_id for unit in report.units}
    result: set[CorpusEvaluationKey] = set()
    for finding in report.findings:
        if finding.unit_id not in unit_ids:
            continue
        result.add(CorpusEvaluationKey(project_id=finding.unit_id, criterion=finding.criterion.value))
    return result


def _header_map(sheet: Any) -> dict[str, int]:
    """Возвращает индексы колонок по первой строке Excel."""

    result: dict[str, int] = {}
    for index, cell in enumerate(sheet[1]):
        if cell.value:
            result[str(cell.value).strip()] = index
    return result


def _cell_text(row: tuple[Any, ...], index: int) -> str:
    """Безопасно достаёт текст ячейки."""

    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _match_project(raw_project: str, candidates: list[_ProjectCandidate]) -> tuple[_ProjectCandidate, float]:
    """Сопоставляет короткое имя проекта из Excel с фактической папкой."""

    normalized = _normalize_project_name(raw_project)
    tokens = set(_project_tokens(raw_project))
    best: tuple[_ProjectCandidate, float] | None = None
    for candidate in candidates:
        score = _project_match_score(normalized, tokens, candidate)
        if best is None or score > best[1]:
            best = (candidate, score)
    if best is None:
        raise ValueError("В отчёте нет единиц контента для сопоставления с Excel.")
    return best


def _project_match_score(normalized: str, tokens: set[str], candidate: _ProjectCandidate) -> float:
    """Считает устойчивый score для разных написаний имени проекта."""

    if not normalized:
        return 0.0
    if normalized == candidate.normalized_name:
        return 1.0
    if normalized in candidate.normalized_name or candidate.normalized_name in normalized:
        return 0.95
    token_overlap = len(tokens & set(candidate.tokens)) / max(len(tokens), 1)
    sequence_score = SequenceMatcher(a=normalized, b=candidate.normalized_name).ratio()
    prefix_score = 0.0
    if tokens and candidate.tokens and next(iter(tokens)) in candidate.tokens:
        prefix_score = 0.25
    return max(sequence_score, token_overlap + prefix_score)


def _criteria_from_gold_row(raw_problem: str, details: str) -> list[str]:
    """Выводит наши критерии из свободного описания проблемы."""

    problem_text = raw_problem.lower()
    detail_text = details.lower()
    criteria: list[Criterion] = []

    # Тип проблемы надёжнее деталей, поэтому сначала используем его.
    for criterion, aliases in CRITERION_ALIASES.items():
        if any(alias in problem_text for alias in aliases):
            criteria.append(criterion)

    # Детали добавляют критерии только по сильным маркерам, чтобы не раздувать эталон.
    detail_markers: dict[Criterion, tuple[str, ...]] = {
        Criterion.ACTUALITY: ("сломанная ссылка", "неактуаль", "устар", "старый стандарт", "версии ос"),
        Criterion.CHECKLIST_ALIGNMENT: ("в чеклисте", "в чек-листе", "чеклист", "чек лист"),
        Criterion.READABILITY: ("опечат", "нумерация", "пронумер", "грамматика"),
        Criterion.CORRECTNESS: ("противореч", "некоррект", "по факту", "не является", "отсутствует"),
    }
    for criterion, markers in detail_markers.items():
        if any(marker in detail_text for marker in markers):
            criteria.append(criterion)

    return [criterion.value for criterion in dict.fromkeys(criteria)]


def _per_criterion_metrics(
    gold_keys: set[CorpusEvaluationKey],
    predicted_keys: set[CorpusEvaluationKey],
) -> list[CriterionMetrics]:
    """Считает метрики по каждому критерию, который есть в эталоне или прогнозе."""

    criteria = sorted({item.criterion for item in gold_keys | predicted_keys})
    metrics: list[CriterionMetrics] = []
    for criterion in criteria:
        gold = {item for item in gold_keys if item.criterion == criterion}
        predicted = {item for item in predicted_keys if item.criterion == criterion}
        tp = len(gold & predicted)
        fp = len(predicted - gold)
        fn = len(gold - predicted)
        metrics.append(
            CriterionMetrics(
                criterion=criterion,
                label=_criterion_label(criterion),
                gold_total=len(gold),
                predicted_total=len(predicted),
                true_positive=tp,
                false_positive=fp,
                false_negative=fn,
                precision=_safe_ratio(tp, tp + fp),
                recall=_safe_ratio(tp, tp + fn),
                f1_score=_f1(tp, fp, fn),
            )
        )
    return metrics


def _write_metrics_csv(summary: CorpusEvaluationSummary, output_path: Path) -> None:
    """Пишет метрики по критериям в CSV."""

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "criterion",
                "label",
                "gold_total",
                "predicted_total",
                "true_positive",
                "false_positive",
                "false_negative",
                "precision",
                "recall",
                "f1_score",
            ],
        )
        writer.writeheader()
        for item in summary.per_criterion:
            writer.writerow(item.model_dump(mode="json"))


def _write_items_csv(items: list[CorpusEvaluationKey], output_path: Path) -> None:
    """Пишет ошибки сравнения: ложные пропуски или ложные срабатывания."""

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["project_id", "criterion", "label"])
        writer.writeheader()
        for item in items:
            writer.writerow(
                {
                    "project_id": item.project_id,
                    "criterion": item.criterion,
                    "label": _criterion_label(item.criterion),
                }
            )


def _normalize_project_name(value: str) -> str:
    """Нормализует имя проекта для сопоставления Excel и папок."""

    text = value.lower().replace("с", "c")
    text = re.sub(r"\.id_\d+.*$", "", text)
    text = re.sub(r"\(\d+\)", "", text)
    text = re.sub(r"[^a-zа-яё0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _project_tokens(value: str) -> list[str]:
    """Выделяет значимые токены имени проекта."""

    normalized = _normalize_project_name(value)
    return [token for token in normalized.split() if token not in {"master", "id"}]


def _criterion_label(criterion_value: str) -> str:
    """Возвращает русское название критерия."""

    try:
        return CRITERION_LABELS[Criterion(criterion_value)]
    except ValueError:
        return criterion_value


def _safe_ratio(numerator: int, denominator: int) -> float:
    """Делит без исключения на пустом наборе."""

    return round(numerator / denominator, 4) if denominator else 0.0


def _f1(true_positive: int, false_positive: int, false_negative: int) -> float:
    """Считает F1 через precision и recall."""

    precision = _safe_ratio(true_positive, true_positive + false_positive)
    recall = _safe_ratio(true_positive, true_positive + false_negative)
    return round(2 * precision * recall / (precision + recall), 4) if precision + recall else 0.0


def _mean(values: list[float]) -> float:
    """Среднее значение для macro-метрик."""

    return round(sum(values) / len(values), 4) if values else 0.0
